import logging
import json
import os
import re
import asyncio
import time
from datetime import datetime
from pathlib import Path
import shutil
import pytz

from faker import Faker
import pyotp
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
    PicklePersistence,
)
from telegram.constants import ParseMode

# --- Configuration & Setup ---
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("apscheduler").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# File paths - using absolute paths for Render
BASE_DIR = Path(__file__).parent
CONFIG_FILE = BASE_DIR / "config.json"
DATA_FILE = BASE_DIR / "data.json"
UPLOADS_DIR = BASE_DIR / "uploads"
USER_UPLOADS_DIR = UPLOADS_DIR / "user_files"

# Load configuration from environment variables or config file
BOT_TOKEN = os.environ.get('BOT_TOKEN')
ADMIN_IDS = list(map(int, os.environ.get('ADMIN_IDS', '').split(','))) if os.environ.get('ADMIN_IDS') else None
SUPPORT_USERNAME = os.environ.get('SUPPORT_USERNAME')

if not all([BOT_TOKEN, ADMIN_IDS, SUPPORT_USERNAME]):
    if not CONFIG_FILE.exists():
        logger.error("config.json not found and environment variables not set!")
        exit()
    with open(CONFIG_FILE, "r") as f:
        config = json.load(f)
    BOT_TOKEN = BOT_TOKEN or config.get("BOT_TOKEN")
    ADMIN_IDS = ADMIN_IDS or config.get("ADMIN_IDS")
    SUPPORT_USERNAME = SUPPORT_USERNAME or config.get("SUPPORT_USERNAME", "@shihab98bc")

# Global lock for thread-safe number distribution and cooldown
number_lock = asyncio.Lock()
last_number_time = 0
COOLDOWN_SECONDS = 10

# Scheduler for automated tasks
scheduler = AsyncIOScheduler()

# --- Data Handling Helper Functions ---
def save_data(data):
    """Saves data to data.json"""
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)

def load_data():
    """
    Loads data from data.json, creating/fixing it if it doesn't exist or is corrupt.
    """
    default_data = {
        "buttons": [],
        "users": {},
        "number_progress": {},
        "blacklist": [],
        "user_2fa_secrets": {},
        "otp_group_link": "https://t.me/+p2ppOgkSosNhZGI1",
        "file_submission_buttons": [],
        "delivery_schedule": {
            "time": None,
            "timezone": "Asia/Dhaka",
            "job_id": "daily_merge_job"
        }
    }
    
    data = default_data

    try:
        if not DATA_FILE.exists() or DATA_FILE.stat().st_size == 0:
            logger.info(f"{DATA_FILE} not found or is empty. Creating a new one.")
            save_data(default_data)
            return default_data

        with open(DATA_FILE, "r") as f:
            data = json.load(f)

        fixed = False
        for key, default_value in default_data.items():
            if key not in data or not isinstance(data.get(key), type(default_value)):
                logger.warning(
                    f"Key '{key}' is missing or has the wrong type in {DATA_FILE}. "
                    f"Resetting it to the default value."
                )
                data[key] = default_value
                fixed = True

        if 'buttons' in data and isinstance(data['buttons'], list):
            new_buttons = []
            for item in data['buttons']:
                if isinstance(item, str):
                    logger.info(f"Converting old button '{item}' to new dictionary format.")
                    new_buttons.append({"name": item, "sub_buttons": []})
                    fixed = True
                elif isinstance(item, dict) and 'name' in item:
                    if 'sub_buttons' not in item or not isinstance(item['sub_buttons'], list):
                        item['sub_buttons'] = []
                        fixed = True
                    new_buttons.append(item)
                else:
                    logger.warning(f"Invalid item found in 'buttons' list: {item}. Skipping.")
                    fixed = True
            data['buttons'] = new_buttons
        
        if fixed:
            logger.info(f"Corrected data structure in {DATA_FILE}. Saving changes.")
            save_data(data)
            
        return data

    except json.JSONDecodeError as e:
        logger.error(f"{DATA_FILE} is corrupted and cannot be read ({e}). Creating a fresh one.")
        save_data(default_data)
        return default_data

# --- Helper to escape markdown characters ---
def escape_markdown(text: str) -> str:
    """Escapes special characters for Telegram MarkdownV2."""
    escape_chars = r'_*[]()~`>#+-=|{}.!'
    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', str(text))

# Initial load of data
load_data()

# Ensure uploads directories exist
UPLOADS_DIR.mkdir(exist_ok=True)
USER_UPLOADS_DIR.mkdir(exist_ok=True)

# --- User & Admin Helper Functions ---
def is_admin(user_id: int) -> bool:
    """Checks if a user is an admin."""
    return user_id in ADMIN_IDS

def is_blacklisted(user_id: int) -> bool:
    """Checks if a user is blacklisted."""
    data = load_data()
    return user_id in data.get("blacklist", [])

def get_keyboard(buttons: list, items_per_row: int = 2) -> ReplyKeyboardMarkup:
    """Creates a ReplyKeyboardMarkup from a list of buttons."""
    rows = [buttons[i:i + items_per_row] for i in range(0, len(buttons), items_per_row)]
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)

async def set_user_state(context: ContextTypes.DEFAULT_TYPE, state: str | None):
    """Sets the next_action state for a user."""
    context.user_data['next_action'] = state

# --- Keyboard Menus ---
MAIN_MENU_USER = get_keyboard(["🔢 Get Number", "✍️ Submit File", "🎭 Fake Name", "📲 Get 2FA", "ℹ️ Info", "🆘 Support"])
MAIN_MENU_ADMIN = get_keyboard(["🔢 Get Number", "✍️ Submit File", "🎭 Fake Name", "📲 Get 2FA", "ℹ️ Info", "🆘 Support", "⚙️ Admin Panel", "📢 Broadcast"])
ADMIN_PANEL_MENU = get_keyboard([
    "➕ Add Button", "🗑️ Remove Button", "📤 Upload File",
    "✍️ Add File Name", "❌ Remove File Name", "⏰ Set time (Bangladesh)",
    "👥 User List", "🔗 Set OTP Group Link", "🚫 Off OTP Group Link",
    "⬅️ Back to Main Menu"
], items_per_row=3)
ADD_REMOVE_BUTTON_MENU = get_keyboard(["1️⃣ Add Main Button", "2️⃣ Add Sub Button", "↩️ Back to Admin Panel"])
REMOVE_BUTTON_MENU = get_keyboard(["1️⃣ Remove Main Button", "2️⃣ Remove Sub Button", "↩️ Back to Admin Panel"])
GENDER_SELECTION_MENU = get_keyboard(["👨 Male", "👩 Female", "⬅️ Back to Main Menu"])
CANCEL_ACTION_MENU = get_keyboard(["↩️ Back to Admin Panel"], items_per_row=1)

# --- Start & Back Commands ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handles the /start command. Welcomes new users once and shows the main menu. For existing users, it just clears state and shows the main menu.
    """
    user = update.effective_user
    if is_blacklisted(user.id):
        await update.message.reply_text(
            "*You have been blocked from using this bot\\.*",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    data = load_data()
    if str(user.id) not in data["users"]:
        data["users"][str(user.id)] = {
            "first_name": user.first_name,
            "last_name": user.last_name,
            "username": user.username,
        }
        save_data(data)

    await set_user_state(context, None)
    context.user_data.clear()  # Clear all context to prevent stale data
    context.user_data['welcomed'] = context.user_data.get('welcomed', False)

    if not context.user_data['welcomed']:
        welcome_text = escape_markdown(f"👋 Welcome, {user.first_name}!")
        await update.message.reply_text(
            f"*{welcome_text}*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        context.user_data['welcomed'] = True

    reply_markup = MAIN_MENU_ADMIN if is_admin(user.id) else MAIN_MENU_USER
    await update.message.reply_text(
        "*Please choose an option from the main menu:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2,
    )

async def back_to_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Navigates back to the main admin panel."""
    logger.info(f"User {update.effective_user.id} returning to admin panel.")
    await set_user_state(context, None)
    context.user_data.clear()  # Clear context to avoid conflicts
    await update.message.reply_text(
        "*⚙️ Admin Panel:*",
        reply_markup=ADMIN_PANEL_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

# --- Main Feature Handlers ---
async def get_number_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Shows the main buttons as a keyboard menu."""
    data = load_data()
    buttons = [btn['name'] for btn in data.get("buttons", [])]
    if not buttons:
        await update.message.reply_text(
            "*Sorry, no number categories are available right now\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    keyboard_buttons = buttons + ["⬅️ Back to Main Menu"]
    reply_markup = get_keyboard(keyboard_buttons)
    await set_user_state(context, 'awaiting_main_category')
    await update.message.reply_text(
        "*Please choose a category to get a number:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def show_sub_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Shows the sub-buttons for a selected main button."""
    main_button_name = update.message.text
    data = load_data()
    
    main_button = next((btn for btn in data['buttons'] if btn['name'] == main_button_name), None)
    
    if not main_button:
        await update.message.reply_text(
            "*Invalid category\\. Please select from the menu\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await get_number_menu(update, context)
        return

    sub_buttons = main_button.get('sub_buttons', [])
    if not sub_buttons:
        await update.message.reply_text(
            "*Sorry, no numbers are available in this sub\\-category yet\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await get_number_menu(update, context)
        return

    keyboard_buttons = sub_buttons + ["⬅️ Back to Main Menu"]
    reply_markup = get_keyboard(keyboard_buttons)
    await set_user_state(context, 'awaiting_number_category')
    context.user_data['main_button_context'] = main_button_name
    await update.message.reply_text(
        f"*Please choose a sub\\-category from '{escape_markdown(main_button_name)}':*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def give_number(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the logic for giving a number to a user."""
    global last_number_time
    sub_button_name = update.message.text
    main_button_name = context.user_data.get('main_button_context')

    if not main_button_name:
        await update.message.reply_text(
            "*An error occurred\\. Please try again from the main menu\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await start(update, context)
        return

    async with number_lock:
        current_time = asyncio.get_event_loop().time()
        if current_time < last_number_time + COOLDOWN_SECONDS:
            await update.message.reply_text(
                f"*Please wait {COOLDOWN_SECONDS} seconds before taking another number\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )
            return

        file_path = UPLOADS_DIR / f"{main_button_name}_{sub_button_name}.txt"
        if not file_path.exists():
            await update.message.reply_text(
                "*Sorry, the numbers for this category have not been uploaded yet\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )
            return

        data = load_data()
        
        if main_button_name not in data.get("number_progress", {}):
            data["number_progress"][main_button_name] = {}
        line_index = data.get("number_progress", {}).get(main_button_name, {}).get(sub_button_name, 0)
        otp_group_link = data.get("otp_group_link", "")

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            if line_index < len(lines):
                number = lines[line_index].strip()
                
                message_text = f"✅ আপনার নাম্বার \\- `{escape_markdown(number)}`\n\n"
                if otp_group_link:
                    message_text += (
                        "এই নাম্বারের Code রিসিভ করার জন্য নিচে Click Here এ Click করুন\\!\n"
                        f"OTP Group \\- [Click here]({escape_markdown(otp_group_link)})\n\n"
                    )
                message_text += "⚙️যেকোনো সমস্যা হলে নিচের Support বাটনে ক্লিক করে আমাদের জানান।"
                
                await update.message.reply_text(
                    message_text,
                    parse_mode=ParseMode.MARKDOWN_V2,
                    disable_web_page_preview=True
                )
                
                data["number_progress"][main_button_name][sub_button_name] = line_index + 1
                save_data(data)
                last_number_time = current_time
            else:
                await update.message.reply_text(
                    "*Sorry, all numbers for this category have been distributed\\.*",
                    parse_mode=ParseMode.MARKDOWN_V2
                )
        except Exception as e:
            logger.error(f"Error reading number file {file_path}: {e}")
            await update.message.reply_text(
                "*An error occurred\\. Please try again later\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )

# --- User File Submission ---
async def submit_file_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows the dynamic list of sub-buttons for file submission."""
    data = load_data()
    buttons = data.get("file_submission_buttons", [])
    if not buttons:
        await update.message.reply_text(
            "*Sorry, no file submission categories are available right now\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    keyboard_buttons = buttons + ["⬅️ Back to Main Menu"]
    reply_markup = get_keyboard(keyboard_buttons)
    await set_user_state(context, 'awaiting_file_submission_category')
    await update.message.reply_text(
        "*Please choose a category to submit your file:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def handle_file_submission_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """After user selects a category, asks for the .xlsx file."""
    category_name = update.message.text
    data = load_data()

    if category_name not in data.get("file_submission_buttons", []):
        await update.message.reply_text("*Invalid category selection\\.*", parse_mode=ParseMode.MARKDOWN_V2)
        await submit_file_menu(update, context)
        return

    context.user_data['file_submission_category'] = category_name
    await set_user_state(context, 'awaiting_xlsx_upload')
    
    message = (
        f"*You have selected '{escape_markdown(category_name)}'\\.*\n\n"
        "*Please upload your \\.xlsx file now\\.*"
    )
    
    await update.message.reply_text(
        message,
        reply_markup=get_keyboard(["⬅️ Back to Main Menu"]),
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def receive_user_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Receives and saves a user's .xlsx file."""
    document = update.message.document
    user_id = update.effective_user.id
    category = context.user_data.get('file_submission_category')

    if not category:
        await update.message.reply_text(
            "*An error occurred, please select a category first\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    # Create user-specific directory if it doesn't exist
    user_dir = USER_UPLOADS_DIR / str(user_id)
    user_dir.mkdir(exist_ok=True)
    
    file_path = user_dir / f"{category}.xlsx"
    
    # Remove old file if it exists
    if file_path.exists():
        os.remove(file_path)
        logger.info(f"Replacing existing file for user {user_id} in category '{category}'.")

    new_file = await document.get_file()
    await new_file.download_to_drive(file_path)

    await update.message.reply_text(
        f"*✅ Your file for '{escape_markdown(category)}' has been successfully submitted\\!*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    # Go back to the main menu
    await start(update, context)

async def fake_name_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Asks the user to select a gender for the fake name."""
    await set_user_state(context, 'awaiting_gender_for_fakename')
    await update.message.reply_text(
        "*Please select a gender for the fake name:*",
        reply_markup=GENDER_SELECTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def fake_name_generate(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Generates and sends a fake identity based on gender."""
    gender_choice = update.message.text
    fake = Faker()

    if gender_choice == "👨 Male":
        first_name = fake.first_name_male()
        gender_emoji = "👨"
    elif gender_choice == "👩 Female":
        first_name = fake.first_name_female()
        gender_emoji = "👩"
    else:
        await update.message.reply_text("*Invalid choice\\. Please try again\\.*", parse_mode=ParseMode.MARKDOWN_V2)
        return

    last_name = fake.last_name()
    username = f"{first_name.lower().replace(' ', '')}{last_name.lower().replace(' ', '')}{fake.random_int(10, 99)}"
    password = fake.password(length=12, special_chars=True, digits=True, upper_case=True, lower_case=True)
    date_str = datetime.now().strftime("%d")
    final_password = f"{password}{date_str}"

    message = (
        f"*{gender_emoji} Generated Identity:*\n\n"
        f"*First name:* `{escape_markdown(first_name)}`\n"
        f"*Last name:* `{escape_markdown(last_name)}`\n"
        f"*Username:* `{escape_markdown(username)}`\n"
        f"*Password:* `{escape_markdown(final_password)}`"
    )
    
    await update.message.reply_text(message, parse_mode=ParseMode.MARKDOWN_V2)
    await start(update, context)

async def get_2fa_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Starts the 2FA code generation process."""
    await set_user_state(context, 'awaiting_2fa_secret')
    
    data = load_data()
    user_id = str(update.effective_user.id)
    saved_secret = data.get("user_2fa_secrets", {}).get(user_id)
    
    if saved_secret:
        message = (
            "*📲 2FA Code Generator*\n\n"
            "You have a saved 2FA secret key\\. Would you like to\\:\n"
            "1\\. Use the saved key\n"
            "2\\. Enter a new key\n\n"
            "Or send your new 2FA secret key now \\(e\\.g\\., BK5V TVQ7 D2RB\\.\\.\\.\\)"
        )
        reply_markup = get_keyboard(["Use saved key", "Enter new key", "⬅️ Back to Main Menu"])
    else:
        message = (
            "*📲 2FA Code Generator*\n\n"
            "Please enter your 2FA secret key \\(e\\.g\\., BK5V TVQ7 D2RB\\.\\.\\.\\)\n\n"
            "*Note\\:* This key will be saved for future use unless you choose to remove it\\."
        )
        reply_markup = get_keyboard(["⬅️ Back to Main Menu"])
    
    await update.message.reply_text(
        message,
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def handle_2fa_secret(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 2FA secret input from the user."""
    user_input = update.message.text
    user_id = str(update.effective_user.id)
    data = load_data()
    
    if user_input == "⬅️ Back to Main Menu":
        await start(update, context)
        return
    elif user_input == "Use saved key":
        saved_secret = data.get("user_2fa_secrets", {}).get(user_id)
        if saved_secret:
            await generate_and_send_2fa_code(update, context, saved_secret)
        else:
            await update.message.reply_text(
                "*No saved 2FA key found\\. Please enter a new one\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )
        return
    elif user_input == "Enter new key":
        await update.message.reply_text(
            "*Please enter your new 2FA secret key \\(e\\.g\\., BK5V TVQ7 D2RB\\.\\.\\.\\)\\:*",
            reply_markup=get_keyboard(["⬅️ Back to Main Menu"]),
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return
    
    def generate_totp(secret: str) -> str:
        """Generate a TOTP code from a secret key."""
        try:
            normalized_secret = secret.replace(" ", "").upper()
            totp = pyotp.TOTP(normalized_secret)
            return totp.now()
        except Exception as e:
            logger.error(f"Error generating TOTP: {e}")
            return None

    def is_valid_2fa_secret(secret: str) -> bool:
        """Check if a string is a valid 2FA secret key."""
        try:
            cleaned = secret.replace(" ", "").upper()
            if len(cleaned) < 16:
                return False
            if not re.match(r'^[A-Z2-7]+$', cleaned):
                return False
            return True
        except:
            return False

    if not is_valid_2fa_secret(user_input):
        await update.message.reply_text(
            "*Invalid 2FA secret key format\\. Please enter a valid key \\(e\\.g\\., BK5V TVQ7 D2RB\\.\\.\\.\\)*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return
    
    if "user_2fa_secrets" not in data:
        data["user_2fa_secrets"] = {}
    data["user_2fa_secrets"][user_id] = user_input
    save_data(data)
    
    await generate_and_send_2fa_code(update, context, user_input)

async def generate_and_send_2fa_code(update: Update, context: ContextTypes.DEFAULT_TYPE, secret: str) -> None:
    """Generates and sends the 2FA code to the user."""
    def generate_totp(secret: str) -> str:
        """Generate a TOTP code from a secret key."""
        try:
            normalized_secret = secret.replace(" ", "").upper()
            totp = pyotp.TOTP(normalized_secret)
            return totp.now()
        except Exception as e:
            logger.error(f"Error generating TOTP: {e}")
            return None
    
    code = generate_totp(secret)
    if not code:
        await update.message.reply_text(
            "*Error generating 2FA code\\. Please check your secret key and try again\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return
    
    remaining_time = 30 - (int(time.time()) % 30)
    
    message = (
        "*🔐 2FA Authentication Code*\n\n"
        f"*Your Code\\:* `{code}`\n"
        f"*Valid for\\:* {remaining_time} seconds\n\n"
        "*Note\\:* This code refreshes every 30 seconds\\. You can request a new code at any time\\."
    )
    
    await update.message.reply_text(
        message,
        reply_markup=MAIN_MENU_ADMIN if is_admin(update.effective_user.id) else MAIN_MENU_USER,
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await set_user_state(context, None)

async def info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Displays user information."""
    user = update.effective_user
    user_info = (
        f"*ℹ️ Your Info\\:*\n\n"
        f"*▪️ ID\\:* `{user.id}`\n"
        f"*▪️ First Name\\:* {escape_markdown(user.first_name)}\n"
        f"*▪️ Last Name\\:* {escape_markdown(user.last_name or 'N/A')}\n"
        f"*▪️ Username\\:* @{escape_markdown(user.username or 'N/A')}\n"
        f"*▪️ Language\\:* {escape_markdown(user.language_code or 'N/A')}"
    )
    await update.message.reply_text(user_info, parse_mode=ParseMode.MARKDOWN_V2)

async def support(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Provides a link to the support contact."""
    await update.message.reply_text(
        f"*🆘 For support, please contact\\:* {SUPPORT_USERNAME}",
        parse_mode=ParseMode.MARKDOWN_V2
    )

# --- Admin Panel Logic ---
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Shows the main admin keyboard."""
    await set_user_state(context, None)
    context.user_data.clear()  # Clear context to avoid stale data
    await update.message.reply_text(
        "*⚙️ Admin Panel:*",
        reply_markup=ADMIN_PANEL_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

# --- Admin: Broadcast ---
async def broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Asks the admin to provide the broadcast message."""
    await set_user_state(context, 'awaiting_broadcast_message')
    await update.message.reply_text(
        "*Please send the message you want to broadcast to all users, or go back\\.*",
        reply_markup=CANCEL_ACTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Broadcasts the message to all users."""
    message_text = update.message.text
    if message_text == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    escaped_message = escape_markdown(message_text)
    notification_text = f"*📢 Broadcast Message:*\n\n{escaped_message}"

    for user_id in data.get("users", {}).keys():
        if int(user_id) != update.effective_user.id:
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=notification_text,
                    parse_mode=ParseMode.MARKDOWN_V2
                )
            except Exception as e:
                logger.warning(f"Could not send broadcast to user {user_id}: {e}")

    await update.message.reply_text(
        "*✅ Broadcast message sent to all users\\!*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)

# --- Admin: Add/Remove File Submission Buttons ---
async def add_file_name_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Asks for the name of the new file submission button."""
    await set_user_state(context, 'awaiting_file_name_add')
    await update.message.reply_text(
        "*Please send the name for the new file submission category \\(e\\.g\\., Netflix\\), or go back\\.*",
        reply_markup=CANCEL_ACTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def add_file_name_receive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Adds the new file submission button name."""
    button_name = update.message.text
    if button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    if button_name in data["file_submission_buttons"]:
        await update.message.reply_text(
            "*This category name already exists\\. Please choose another name\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    data["file_submission_buttons"].append(button_name)
    save_data(data)
    await update.message.reply_text(
        f"*✅ File submission category '{escape_markdown(button_name)}' added successfully\\!*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)

async def remove_file_name_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows existing file submission buttons for removal."""
    data = load_data()
    buttons = data.get("file_submission_buttons", [])
    if not buttons:
        await update.message.reply_text(
            "*There are no file submission categories to remove\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    keyboard = get_keyboard(buttons + ["↩️ Back to Admin Panel"])
    await set_user_state(context, 'awaiting_file_name_to_remove')
    await update.message.reply_text(
        "*Select a file submission category to remove:*",
        reply_markup=keyboard,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def remove_file_name_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Removes the selected file submission button."""
    button_name = update.message.text
    if button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    if button_name in data["file_submission_buttons"]:
        data["file_submission_buttons"].remove(button_name)
        save_data(data)
        await update.message.reply_text(
            f"*🗑️ File submission category '{escape_markdown(button_name)}' removed\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    else:
        await update.message.reply_text(
            "*Error: Category not found\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    
    await back_to_admin_panel(update, context)

# --- Admin: Set Delivery Time & Scheduling ---
async def set_delivery_time_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Asks admin for the daily delivery time."""
    data = load_data()
    schedule_info = data.get("delivery_schedule", {})
    current_time_24h = schedule_info.get("time")
    
    current_time_str = "Not set"
    if current_time_24h:
        try:
            time_obj = datetime.strptime(current_time_24h, "%H:%M").time()
            current_time_str = time_obj.strftime("%I:%M %p")
        except ValueError:
            current_time_str = "Invalid format"

    await set_user_state(context, 'awaiting_delivery_time')
    await update.message.reply_text(
        "*Please send the daily delivery time for the master file\\.*\n"
        f"\\(Current: *{escape_markdown(current_time_str)}*\\)\n\n"
        "Use 12\\-hour format with AM/PM \\(e\\.g\\., `10:30 AM`, `7:45 PM`\\)\\.",
        reply_markup=CANCEL_ACTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def set_delivery_time_receive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Receives, validates, and sets the daily delivery time."""
    time_input = update.message.text
    if time_input == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    # Regex to validate "HH:MM AM/PM" format and capture parts
    match = re.match(r'^(0?[1-9]|1[0-2]):([0-5][0-9])\s*([AP]M)$', time_input, re.IGNORECASE)
    
    if not match:
        await update.message.reply_text(
            "*Invalid time format\\.* Please use a valid 12\\-hour format like `10:30 AM` or `7:45 PM`\\.",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    hour_12, minute, am_pm = match.groups()
    hour_12, minute = int(hour_12), int(minute)
    am_pm = am_pm.upper()

    hour_24 = hour_12
    if am_pm == 'PM' and hour_12 != 12:
        hour_24 += 12
    elif am_pm == 'AM' and hour_12 == 12:
        hour_24 = 0

    data = load_data()
    data["delivery_schedule"]["time"] = f"{hour_24:02d}:{minute:02d}"
    save_data(data)
    
    # Reschedule the job
    await schedule_daily_job(context.application)

    await update.message.reply_text(
        f"*✅ Daily delivery time set to {escape_markdown(time_input)} Bangladesh Time\\.*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)

async def merge_and_send_files(application: Application):
    """
    Merges user files for each category into a separate report
    and sends each report individually to admins only if there is data.
    """
    logger.info("Starting daily file merge and send process.")
    data = load_data()
    categories = data.get("file_submission_buttons", [])
    report_date = datetime.now(pytz.timezone('Asia/Dhaka')).strftime("%Y-%m-%d")

    if not categories:
        logger.info("No file submission categories configured. Skipping merge.")
        return

    # Process each category individually
    for category in categories:
        logger.info(f"Processing report for category: {category}")
        
        # Create a new workbook for each category
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = category

        user_files = list(USER_UPLOADS_DIR.glob(f"*/{category}.xlsx"))

        if not user_files:
            logger.warning(f"No files submitted for category '{category}'. Skipping.")
            continue
            
        has_data_for_category = False
        is_header_written = False

        # Aggregate data from all users for the current category
        for user_file_path in user_files:
            try:
                user_wb = openpyxl.load_workbook(user_file_path)
                user_sheet = user_wb.active
                
                # Write header only once from the first file
                if not is_header_written and user_sheet.max_row > 0:
                    header = [cell.value for cell in user_sheet[1] if cell.value is not None]
                    if header:
                        sheet.append(header)
                        is_header_written = True

                # Append data rows (skip header)
                for row_idx in range(2, user_sheet.max_row + 1):
                    row_data = [cell.value for cell in user_sheet[row_idx]]
                    if any(row_data):  # only append if row is not empty
                        sheet.append(row_data)
                        has_data_for_category = True
                        
            except (InvalidFileException, Exception) as e:
                logger.error(f"Could not process file {user_file_path}: {e}")

        # Delete user files after processing
        for user_file_path in user_files:
            os.remove(user_file_path)

        # Additional check to ensure there is data beyond header
        if sheet.max_row <= 1 or not has_data_for_category:
            logger.info(f"Only header or no data for category '{category}'. Skipping report.")
            continue
        
        report_filename = f"{category}_Report_{report_date}.xlsx"
        wb.save(report_filename)
        
        caption = f"*Daily User File Report for '{escape_markdown(category)}' on {escape_markdown(report_date)}*"
        
        for admin_id in ADMIN_IDS:
            try:
                await application.bot.send_document(
                    chat_id=admin_id,
                    document=open(report_filename, 'rb'),
                    filename=report_filename,
                    caption=caption,
                    parse_mode=ParseMode.MARKDOWN_V2
                )
                logger.info(f"Sent '{category}' report to admin {admin_id}.")
            except Exception as e:
                logger.error(f"Failed to send '{category}' report to admin {admin_id}: {e}")
                
        # Clean up the generated report file
        os.remove(report_filename)


async def schedule_daily_job(application: Application):
    """Schedules or reschedules the daily file merge job based on data.json."""
    data = load_data()
    schedule_info = data["delivery_schedule"]
    job_id = schedule_info["job_id"]
    
    # Remove existing job before adding a new one
    if scheduler.get_job(job_id):
        scheduler.remove_job(job_id)
        logger.info(f"Removed existing schedule job '{job_id}'.")

    if schedule_info["time"]:
        try:
            hour, minute = map(int, schedule_info["time"].split(':'))
            timezone = pytz.timezone(schedule_info["timezone"])
            
            scheduler.add_job(
                merge_and_send_files,
                trigger=CronTrigger(hour=hour, minute=minute, timezone=timezone),
                id=job_id,
                name="Daily File Merge",
                args=[application]
            )
            logger.info(f"Scheduled daily report for {hour:02d}:{minute:02d} in {timezone}.")
        except Exception as e:
            logger.error(f"Failed to schedule daily job: {e}")

# --- Admin: Add/Remove Number Buttons (Original Logic) ---

async def add_button_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows options to add a main or sub button."""
    await set_user_state(context, 'awaiting_add_type')
    await update.message.reply_text(
        "*Select the type of button you want to add:*",
        reply_markup=ADD_REMOVE_BUTTON_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def add_main_button_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Asks for a new main button name."""
    await set_user_state(context, 'awaiting_main_button_name')
    await update.message.reply_text(
        "*Please send the name for the new main button \\(e\\.g\\., OTT\\), or go back\\.*",
        reply_markup=CANCEL_ACTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def add_main_button_receive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Adds the new main button to the data structure."""
    button_name = update.message.text
    if button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    if any(btn['name'] == button_name for btn in data["buttons"]):
        await update.message.reply_text(
            "*This main button already exists\\. Please choose another name\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    data["buttons"].append({"name": button_name, "sub_buttons": []})
    save_data(data)
    await update.message.reply_text(
        f"*✅ Main button '{escape_markdown(button_name)}' added successfully\\!*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)

async def add_sub_button_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows main buttons to select where to add a sub-button."""
    data = load_data()
    main_buttons = [btn['name'] for btn in data.get("buttons", [])]
    if not main_buttons:
        await update.message.reply_text(
            "*Please add a main button first before adding a sub\\-button\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    keyboard_buttons = main_buttons + ["↩️ Back to Admin Panel"]
    reply_markup = get_keyboard(keyboard_buttons)
    await set_user_state(context, 'awaiting_main_for_sub_add')
    await update.message.reply_text(
        "*Select a main button to add a sub\\-button to:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def add_sub_button_ask_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Asks for the name of the new sub-button."""
    main_button_name = update.message.text
    if main_button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    main_button_obj = next((btn for btn in data['buttons'] if btn['name'] == main_button_name), None)

    if not main_button_obj:
        await update.message.reply_text(
            "*Invalid main button selected\\. Please try again\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await add_sub_button_start(update, context)
        return
    
    context.user_data['main_button_context'] = main_button_name
    await set_user_state(context, 'awaiting_sub_button_name')
    await update.message.reply_text(
        f"*Please send the name for the new sub\\-button under '{escape_markdown(main_button_name)}' \\(e\\.g\\., Netflix\\), or go back\\.*",
        reply_markup=CANCEL_ACTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def add_sub_button_receive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Adds the new sub-button to the data structure."""
    sub_button_name = update.message.text
    if sub_button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    main_button_name = context.user_data.get('main_button_context')
    if not main_button_name:
        await update.message.reply_text(
            "*An error occurred\\. Please start over\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    main_button_obj = next((btn for btn in data['buttons'] if btn['name'] == main_button_name), None)

    if not main_button_obj:
        await update.message.reply_text(
            "*An error occurred\\. Please start over\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    if sub_button_name in main_button_obj['sub_buttons']:
        await update.message.reply_text(
            f"*This sub\\-button already exists under '{escape_markdown(main_button_name)}'\\. Please choose another name\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    main_button_obj['sub_buttons'].append(sub_button_name)
    save_data(data)
    await update.message.reply_text(
        f"*✅ Sub\\-button '{escape_markdown(sub_button_name)}' added successfully to '{escape_markdown(main_button_name)}'\\!*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)

async def remove_button_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Asks admin whether to remove a main or sub button."""
    data = load_data()
    if not data.get("buttons", []):
        await update.message.reply_text(
            "*There are no buttons to remove\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return
    
    await set_user_state(context, 'awaiting_remove_type')
    await update.message.reply_text(
        "*Select the type of button you want to remove:*",
        reply_markup=REMOVE_BUTTON_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def remove_main_button_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows main buttons for removal."""
    data = load_data()
    main_buttons = [btn['name'] for btn in data.get("buttons", [])]
    
    keyboard_buttons = main_buttons + ["↩️ Back to Admin Panel"]
    reply_markup = get_keyboard(keyboard_buttons)
    await set_user_state(context, 'awaiting_main_button_to_remove')
    await update.message.reply_text(
        "*Select a main button to remove \\(this will also remove all its sub\\-buttons and files\\):*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def remove_main_button_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Removes the selected main button and its sub-buttons/files."""
    button_name = update.message.text
    if button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    main_button_obj = next((btn for btn in data['buttons'] if btn['name'] == button_name), None)

    if not main_button_obj:
        await update.message.reply_text(
            "*Error\\: Main button not found\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return
    
    for sub_btn in main_button_obj.get('sub_buttons', []):
        file_path = UPLOADS_DIR / f"{button_name}_{sub_btn}.txt"
        if file_path.exists():
            os.remove(file_path)
    
    data.get("number_progress", {}).pop(button_name, None)
    data["buttons"].remove(main_button_obj)
    save_data(data)

    await update.message.reply_text(
        f"*🗑️ Main button '{escape_markdown(button_name)}' and all its sub\\-buttons/files removed\\.*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)

async def remove_sub_button_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows main buttons to select which sub-button to remove."""
    data = load_data()
    main_buttons = [btn['name'] for btn in data.get("buttons", [])]
    
    keyboard_buttons = main_buttons + ["↩️ Back to Admin Panel"]
    reply_markup = get_keyboard(keyboard_buttons)
    await set_user_state(context, 'awaiting_main_for_sub_remove')
    await update.message.reply_text(
        "*Select a main button to see its sub\\-buttons for removal:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def remove_sub_button_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows sub-buttons for removal."""
    main_button_name = update.message.text
    if main_button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    main_button_obj = next((btn for btn in data['buttons'] if btn['name'] == main_button_name), None)

    if not main_button_obj:
        await update.message.reply_text(
            "*Invalid main button selected\\. Please try again\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await remove_sub_button_start(update, context)
        return
    
    sub_buttons = main_button_obj.get('sub_buttons', [])
    if not sub_buttons:
        await update.message.reply_text(
            f"*There are no sub\\-buttons to remove under '{escape_markdown(main_button_name)}'\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return
    
    keyboard_buttons = sub_buttons + ["↩️ Back to Admin Panel"]
    reply_markup = get_keyboard(keyboard_buttons)
    context.user_data['main_button_context'] = main_button_name
    await set_user_state(context, 'awaiting_sub_button_to_remove')
    await update.message.reply_text(
        f"*Select a sub\\-button to remove from '{escape_markdown(main_button_name)}':*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def remove_sub_button_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Removes the selected sub-button and its file."""
    sub_button_name = update.message.text
    if sub_button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    main_button_name = context.user_data.get('main_button_context')
    if not main_button_name:
        await update.message.reply_text(
            "*An error occurred\\. Please start over\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    main_button_obj = next((btn for btn in data['buttons'] if btn['name'] == main_button_name), None)

    if main_button_obj and sub_button_name in main_button_obj['sub_buttons']:
        main_button_obj['sub_buttons'].remove(sub_button_name)
        
        file_path = UPLOADS_DIR / f"{main_button_name}_{sub_button_name}.txt"
        if file_path.exists():
            os.remove(file_path)
        if main_button_name in data.get("number_progress", {}) and sub_button_name in data["number_progress"][main_button_name]:
            data["number_progress"][main_button_name].pop(sub_button_name, None)
        
        save_data(data)
        await update.message.reply_text(
            f"*🗑️ Sub\\-button '{escape_markdown(sub_button_name)}' removed\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    else:
        await update.message.reply_text(
            "*Error\\: Sub\\-button not found\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    await back_to_admin_panel(update, context)

# --- Admin: Upload File ---
async def upload_file_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows main buttons to select where to upload a file."""
    data = load_data()
    main_buttons = [btn['name'] for btn in data.get("buttons", [])]
    
    if not main_buttons:
        await update.message.reply_text(
            "*Please add a main button first before uploading a file\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return
        
    keyboard_buttons = main_buttons + ["↩️ Back to Admin Panel"]
    reply_markup = get_keyboard(keyboard_buttons)
    await set_user_state(context, 'awaiting_main_for_upload')
    await update.message.reply_text(
        "*Select the main button for which you want to upload a file:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def upload_file_select_sub(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows sub-buttons for file upload."""
    main_button_name = update.message.text
    if main_button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    main_button_obj = next((btn for btn in data['buttons'] if btn['name'] == main_button_name), None)

    if not main_button_obj:
        await update.message.reply_text(
            "*Invalid main button selected\\. Please try again\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await upload_file_menu(update, context)
        return
    
    sub_buttons = main_button_obj.get('sub_buttons', [])
    if not sub_buttons:
        await update.message.reply_text(
            f"*Please add a sub\\-button to '{escape_markdown(main_button_name)}' first\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return
        
    keyboard_buttons = sub_buttons + ["↩️ Back to Admin Panel"]
    reply_markup = get_keyboard(keyboard_buttons)
    context.user_data['main_button_context'] = main_button_name
    await set_user_state(context, 'awaiting_upload_button_choice')
    await update.message.reply_text(
        f"*Select the sub\\-button for which you want to upload a \\.txt file\\:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def upload_file_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Asks for the .txt file."""
    sub_button_name = update.message.text
    if sub_button_name == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    main_button_name = context.user_data.get('main_button_context')
    if not main_button_name:
        await update.message.reply_text(
            "*An error occurred\\. Please start over\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    main_button_obj = next((btn for btn in data['buttons'] if btn['name'] == main_button_name), None)
    if not main_button_obj or sub_button_name not in main_button_obj['sub_buttons']:
        await update.message.reply_text(
            "*Invalid sub\\-button selected\\. Please try again\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await upload_file_menu(update, context)
        return

    context.user_data['sub_button_context'] = sub_button_name
    await set_user_state(context, 'awaiting_txt_file_upload')
    
    escaped_main = escape_markdown(main_button_name)
    escaped_sub = escape_markdown(sub_button_name)
    message = f"*Please upload the \\.txt file for the '{escaped_main} \\> {escaped_sub}' button, or go back\\.*"
    
    await update.message.reply_text(
        message,
        reply_markup=CANCEL_ACTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def upload_txt_file_receive(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Receives and processes the uploaded .txt file."""
    document = update.message.document
    main_button_name = context.user_data.get('main_button_context')
    sub_button_name = context.user_data.get('sub_button_context')
    
    if not main_button_name or not sub_button_name:
        await update.message.reply_text(
            "*An error occurred\\. Please start over\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    file_path = UPLOADS_DIR / f"{main_button_name}_{sub_button_name}.txt"
    if file_path.exists():
        os.remove(file_path)
        logger.info(f"Deleted old file for button '{main_button_name}_{sub_button_name}'.")

    new_file = await document.get_file()
    await new_file.download_to_drive(file_path)

    data = load_data()
    if main_button_name not in data.get("number_progress", {}):
        data["number_progress"][main_button_name] = {}
    data['number_progress'][main_button_name][sub_button_name] = 0
    save_data(data)
    
    escaped_main = escape_markdown(main_button_name)
    escaped_sub = escape_markdown(sub_button_name)
    await update.message.reply_text(
        f"*✅ File for '{escaped_main} \\> {escaped_sub}' uploaded successfully\\!*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    
    notification_text = f"*অ্যাডমিন {escaped_main} এর জন্য '{escaped_sub}' এর নতুন নাম্বার যোগ করেছেন। এখন আপনি নাম্বার নিতে পারেন\\!*"

    for user_id in data.get("users", {}).keys():
        if int(user_id) != update.effective_user.id:
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=notification_text,
                    parse_mode=ParseMode.MARKDOWN_V2
                )
            except Exception as e:
                logger.warning(f"Could not notify user {user_id}: {e}")
    
    await back_to_admin_panel(update, context)

# --- Admin: User List ---
async def user_list_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows the list of users for management."""
    data = load_data()
    users = data.get("users", {})
    blacklist = data.get("blacklist", [])
    if not users:
        await update.message.reply_text(
            "*No users have interacted with the bot yet\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    buttons = []
    for user_id, user_info in users.items():
        name = user_info.get('first_name', f"User {user_id}")
        status = " \\(🚫 Blocked\\)" if int(user_id) in blacklist else ""
        buttons.append(f"👤 {name} \\(ID: {user_id}\\){status}")

    buttons.append("↩️ Back to Admin Panel")
    reply_markup = get_keyboard(buttons, items_per_row=1)
    await set_user_state(context, 'awaiting_user_to_manage')
    await update.message.reply_text(
        "*Select a user to manage:*",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def user_manage_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows Block/Unblock options for a selected user."""
    text = update.message.text
    if text == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    match = re.search(r'\(ID:(\d+)\)', text)
    if not match:
        await update.message.reply_text(
            "*Invalid selection\\. Please try again\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await user_list_menu(update, context)
        return

    user_id_to_manage = int(match.group(1))
    context.user_data['user_to_manage'] = user_id_to_manage
    
    keyboard = get_keyboard(["🚫 Block User", "✅ Unblock User", "↩️ Back to Admin Panel"])
    await set_user_state(context, 'awaiting_user_manage_action')
    await update.message.reply_text(
        f"*Managing user `{user_id_to_manage}`\\. Select an action\\:*",
        reply_markup=keyboard,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def user_toggle_block(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Toggles block/unblock status for a user."""
    user_id = context.user_data.get('user_to_manage')
    if not user_id:
        await update.message.reply_text(
            "*Error\\: No user selected\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        await back_to_admin_panel(update, context)
        return

    data = load_data()
    blacklist = data.get("blacklist", [])
    action = update.message.text

    if action == "🚫 Block User":
        if user_id not in blacklist:
            blacklist.append(user_id)
            data['blacklist'] = blacklist
            save_data(data)
            await update.message.reply_text(
                f"*User `{user_id}` has been blocked\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text="*You have been blocked from using this bot\\.*",
                    parse_mode=ParseMode.MARKDOWN_V2
                )
            except Exception as e:
                logger.warning(f"Could not notify user {user_id} of block: {e}")
        else:
            await update.message.reply_text(
                f"*User `{user_id}` is already blocked\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )
    elif action == "✅ Unblock User":
        if user_id in blacklist:
            blacklist.remove(user_id)
            data['blacklist'] = blacklist
            save_data(data)
            await update.message.reply_text(
                f"*User `{user_id}` has been unblocked\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text="*You have been unblocked and can now use the bot again\\.*",
                    parse_mode=ParseMode.MARKDOWN_V2
                )
            except Exception as e:
                logger.warning(f"Could not notify user {user_id} of unblock: {e}")
        else:
            await update.message.reply_text(
                f"*User `{user_id}` is not blocked\\.*",
                parse_mode=ParseMode.MARKDOWN_V2
            )
    else:
        await update.message.reply_text(
            "*Invalid action selected\\. Please try again\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    
    context.user_data.pop('user_to_manage', None)
    await user_list_menu(update, context)

# --- Admin: Set OTP Group Link ---
async def set_otp_link_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Asks the admin to provide the new OTP group link."""
    await set_user_state(context, 'awaiting_otp_link')
    current_link = load_data().get("otp_group_link", "Not set")
    await update.message.reply_text(
        f"*Please send the new OTP group link \\(current\\: {escape_markdown(current_link)}\\), or go back\\.*",
        reply_markup=CANCEL_ACTION_MENU,
        parse_mode=ParseMode.MARKDOWN_V2
    )

async def set_otp_link_receive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Receives and saves the new OTP group link."""
    new_link = update.message.text
    if new_link == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return
    
    if not re.match(r'https?://(?:www\.)?\S+', new_link):
        await update.message.reply_text(
            "*Invalid link format\\. Please send a valid URL starting with http:// or https://\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    data = load_data()
    data["otp_group_link"] = new_link
    save_data(data)
    await update.message.reply_text(
        f"*✅ OTP Group Link updated to\\: {escape_markdown(new_link)}\\!*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)
    
# --- Admin: Off OTP Group Link ---
async def off_otp_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sets the OTP group link to an empty string."""
    data = load_data()
    data["otp_group_link"] = ""
    save_data(data)
    await update.message.reply_text(
        "*✅ OTP Group Link has been turned off\\. It will no longer be shown to users\\.*",
        parse_mode=ParseMode.MARKDOWN_V2
    )
    await back_to_admin_panel(update, context)

# --- Main Handler for Text and State Machine ---
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """The main router for all text messages."""
    user = update.effective_user
    text = update.message.text

    if is_blacklisted(user.id): 
        await update.message.reply_text(
            "*You have been blocked from using this bot\\.*",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    state = context.user_data.get('next_action')
    
    if text == "⬅️ Back to Main Menu":
        await start(update, context)
        return

    if is_admin(user.id) and text == "↩️ Back to Admin Panel":
        await back_to_admin_panel(update, context)
        return

    # State machine routing
    if state:
        # User states
        if state == 'awaiting_main_category': await show_sub_buttons(update, context)
        elif state == 'awaiting_number_category': await give_number(update, context)
        elif state == 'awaiting_gender_for_fakename':
            if text in ["👨 Male", "👩 Female"]: await fake_name_generate(update, context)
        elif state == 'awaiting_2fa_secret': await handle_2fa_secret(update, context)
        elif state == 'awaiting_file_submission_category': await handle_file_submission_category(update, context)
        
        # Admin states
        elif is_admin(user.id):
            if state == 'awaiting_broadcast_message': await broadcast_message(update, context)
            elif state == 'awaiting_add_type':
                if text == "1️⃣ Add Main Button": await add_main_button_start(update, context)
                elif text == "2️⃣ Add Sub Button": await add_sub_button_start(update, context)
            elif state == 'awaiting_remove_type':
                if text == "1️⃣ Remove Main Button": await remove_main_button_start(update, context)
                elif text == "2️⃣ Remove Sub Button": await remove_sub_button_start(update, context)
            elif state == 'awaiting_main_button_name': await add_main_button_receive(update, context)
            elif state == 'awaiting_main_for_sub_add': await add_sub_button_ask_name(update, context)
            elif state == 'awaiting_sub_button_name': await add_sub_button_receive(update, context)
            elif state == 'awaiting_main_button_to_remove': await remove_main_button_action(update, context)
            elif state == 'awaiting_main_for_sub_remove': await remove_sub_button_menu(update, context)
            elif state == 'awaiting_sub_button_to_remove': await remove_sub_button_action(update, context)
            elif state == 'awaiting_main_for_upload': await upload_file_select_sub(update, context)
            elif state == 'awaiting_upload_button_choice': await upload_file_ask(update, context)
            elif state == 'awaiting_txt_file_upload':
                await update.message.reply_text("*Please upload a \\.txt file or use the 'Back' button\\.*", parse_mode=ParseMode.MARKDOWN_V2)
            elif state == 'awaiting_xlsx_upload':
                await update.message.reply_text("*Please upload an \\.xlsx file or use the 'Back' button\\.*", parse_mode=ParseMode.MARKDOWN_V2)
            elif state == 'awaiting_user_to_manage': await user_manage_menu(update, context)
            elif state == 'awaiting_user_manage_action':
                if text in ["🚫 Block User", "✅ Unblock User"]: await user_toggle_block(update, context)
            elif state == 'awaiting_otp_link': await set_otp_link_receive(update, context)
            elif state == 'awaiting_file_name_add': await add_file_name_receive(update, context)
            elif state == 'awaiting_file_name_to_remove': await remove_file_name_action(update, context)
            elif state == 'awaiting_delivery_time': await set_delivery_time_receive(update, context)
        return

    # Standard command routing
    if text == "🔢 Get Number": await get_number_menu(update, context)
    elif text == "✍️ Submit File": await submit_file_menu(update, context)
    elif text == "🎭 Fake Name": await fake_name_start(update, context)
    elif text == "📲 Get 2FA": await get_2fa_start(update, context)
    elif text == "ℹ️ Info": await info(update, context)
    elif text == "🆘 Support": await support(update, context)
    elif is_admin(user.id):
        if text == "⚙️ Admin Panel": await admin_panel(update, context)
        elif text == "➕ Add Button": await add_button_menu(update, context)
        elif text == "🗑️ Remove Button": await remove_button_menu(update, context)
        elif text == "📤 Upload File": await upload_file_menu(update, context)
        elif text == "👥 User List": await user_list_menu(update, context)
        elif text == "🔗 Set OTP Group Link": await set_otp_link_start(update, context)
        elif text == "🚫 Off OTP Group Link": await off_otp_link(update, context)
        elif text == "📢 Broadcast": await broadcast_start(update, context)
        elif text == "✍️ Add File Name": await add_file_name_start(update, context)
        elif text == "❌ Remove File Name": await remove_file_name_start(update, context)
        elif text == "⏰ Set time (Bangladesh)": await set_delivery_time_start(update, context)

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles file uploads, routing them based on user state."""
    user = update.effective_user
    if is_blacklisted(user.id):
        await update.message.reply_text("*You have been blocked from using this bot\\.*", parse_mode=ParseMode.MARKDOWN_V2)
        return
        
    state = context.user_data.get('next_action')
    doc = update.message.document

    # Admin uploading .txt file for numbers
    if is_admin(user.id) and state == 'awaiting_txt_file_upload':
        if not doc.file_name.lower().endswith('.txt'):
            await update.message.reply_text("*Invalid file type\\. Please upload a \\.txt file\\.*", parse_mode=ParseMode.MARKDOWN_V2)
            return
        await upload_txt_file_receive(update, context)
    
    # User uploading .xlsx file for submission
    elif state == 'awaiting_xlsx_upload':
        if not doc.file_name.lower().endswith('.xlsx'):
            await update.message.reply_text("*Invalid file type\\. Please upload an \\.xlsx file\\.*", parse_mode=ParseMode.MARKDOWN_V2)
            return
        await receive_user_file(update, context)
        
    else:
        # Unauthorized or unexpected file upload
        if is_admin(user.id):
             await update.message.reply_text("*I'm not expecting a file right now\\. Please use the menu buttons\\.*", parse_mode=ParseMode.MARKDOWN_V2)
        else:
             await update.message.reply_text("*You are not authorized to perform this action or the bot is not expecting a file\\.*", parse_mode=ParseMode.MARKDOWN_V2)

async def post_init(application: Application) -> None:
    """Function to run after the bot is initialized."""
    await schedule_daily_job(application)
    scheduler.start()
    logger.info("Scheduler started.")


def main() -> None:
    """Start the bot."""
    persistence = PicklePersistence(filepath="bot_data.pkl")
    application = Application.builder().token(BOT_TOKEN).persistence(persistence).post_init(post_init).build()
    
    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()